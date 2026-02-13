from odoo import api, fields, models, _
from odoo.exceptions import ValidationError

import base64
from io import StringIO
import logging
import pathlib
import os
from datetime import datetime
_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Border, Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')



class KSWarehouseReportValuation(models.Model):
    _name = "ks.warehouse.report.valuation"
    _description = "Stock Valuation report"
    # _auto = False

    ks_report = {'product_code':0, 'product_type':1, 'product_categ_id':2, 'product_name':3,
                 'company_id':4, 'product_sales_price':6, 'product_qty_available':7, 'product_id':8, 'product_barcode':9}
    kr_in_dates = {'product_id':0, 'company_id':1, 'opening_stock':2,'qty_in':5,'qty_out':6, 'closing_stock':3, 'qty_date':4, 'opening_value': 7,
                   'closing_value': 8, 'value_date': 9, 'value_in': 10, 'value_out': 11}

    ks_name = fields.Char(default='Stock Valuation Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_inventory_loss = fields.Boolean('Include Inventory Loss')
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda self: self.env.user.company_id)
    ks_show_exhausted = fields.Boolean('Show Exhausted')
    ks_show_opening = fields.Boolean('Show Opening', default=True)
    ks_show_closing = fields.Boolean('Show Closing', default=True)
    ks_show_adjustment = fields.Boolean('Show Adjustment', default=True)
    ks_show_scrap_loss = fields.Boolean('Show Scrap/Loss', default=True)
    ks_show_current = fields.Boolean('Show Current', default=True)

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)


    def ks_generate_xlsx_report(self):
        return self.env.ref('ks_warehouse_report.ks_action_warehouse_report_valuation_action').report_action(self)
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_create_workbook_header(report_name, sheet)

        # get qty available
        if self.ks_inventory_loss:
            self.env.cr.execute("""
            SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
                   ks_product_sales_price, ks_product_qty_available,ks_product_id,ks_product_barcode
            FROM ks_warehouse_report 
            WHERE ks_company_id = %s and ks_product_qty_available != 0 and (ks_usage = 'internal' or ks_usage = 'inventory')
                order by ks_location_id
            """ % self.ks_company_id.id)
        else:
            self.env.cr.execute("""
                    SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
                           ks_product_sales_price, ks_product_qty_available,ks_product_id,ks_product_barcode
                    FROM ks_warehouse_report 
                    WHERE ks_company_id = %s and ks_product_qty_available != 0 and ks_usage = 'internal' 
                        order by ks_location_id
                    """ % self.ks_company_id.id)

        datas = self.env.cr.fetchall()
        if not datas:
            raise ValidationError(_("Opps! There are no data."))

        dates_in = self.ks_data_in_date()

        ks_adjusted_stock = self.ks_adjusted_stock()

        ks_scrap_stock = self.ks_scrap_stock()

        datas = self.ks_merge_data(datas, dates_in, ks_adjusted_stock, ks_scrap_stock)


        if datas:
            i = 1; row = 10; col = 0
            for data in datas:
                # if (not self.location_id or self.location.id == data[4]) or (not self.location_id or self.location.id == data[4])
                sheet.cell(row, 1, i)
                sheet.cell(row, 2, data[0])
                sheet.cell(row, 3, data[19])
                if data[1] == 'product':
                    sheet.cell(row, 4, 'Stockable')
                elif data[1]  == 'consu':
                    sheet.cell(row, 4, 'Consumable')
                if data[2]:
                    catge_id = self.env['product.category'].browse(int(data[2]))
                sheet.cell(row, 5, catge_id.name)
                sheet.cell(row, 6, data[3])
                if data[4]:
                    location_id = self.env['stock.location'].browse(int(data[4]))
                sheet.cell(row, 7, location_id.display_name)
                if data[5]:
                    comp_id = self.env['res.company'].browse(int(data[5]))
                sheet.cell(row, 8, comp_id.name)
                sheet.cell(row, 9, data[6])
                sheet.cell(row, 10, data[7])
                sheet.cell(row, 11, data[8])
                c_1, c_2 = 12, 13
                if self.ks_show_opening:
                    sheet.cell(row, c_1, data[9])
                    sheet.cell(row, c_2, data[10])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_closing:
                    sheet.cell(row, c_1, data[11])
                    sheet.cell(row, c_2, data[12])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_adjustment:
                    sheet.cell(row, c_1, data[13])
                    sheet.cell(row, c_2, data[14])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_scrap_loss:
                    sheet.cell(row, c_1, data[15])
                    sheet.cell(row, c_2, data[16])
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if self.ks_show_current:
                    sheet.cell(row, c_1, data[17])
                    sheet.cell(row, c_2, data[18])

                row += 1
                i += 1
        output = StringIO()
        filename = ('/home/odoo/.local/' + str(report_name) + '.xlsx')
        _log.info("Filepath: %s" % (filename))
        workbook.save(filename)
        fp = open(filename, "rb")
        file_data = fp.read()
        out = base64.encodebytes(file_data)

        # Files actions
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.valuation.out'].create(attach_vals)
        fp.close()
        os.remove(filename)
        _log.info("File closed and removed.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.valuation.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }


    def ks_action_generate_report(self):
        return True




    def ks_scrap_stock(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        self.env.cr.execute("""
            select scrap.product_id, scrap.location_id, sm.company_id, sum(scrap.scrap_qty) 
            from stock_scrap as scrap
                left join stock_move as sm on sm.id = scrap.move_id
            where scrap.state = 'done' and sm.company_id = '%s' and scrap.date_done between '%s' and '%s'
            group by scrap.product_id, scrap.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to)))

        scrap_date = self.env.cr.fetchall()
        if not scrap_date:
            return {}
        else:
            ks_dict = dict()
            for ks in scrap_date: # product_id + location_id + company_id : qty_done(state=done)
                ks_dict[str(ks[0])+str(ks[1])+str(ks[2])] = ks[3]
            scrap_date = ks_dict
        return scrap_date


    def ks_adjusted_stock(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        self.env.cr.execute("""
            select sml.product_id, sml.location_id, sm.company_id, sum(sml.qty_done) 
            from stock_move_line as sml
                left join stock_move as sm on sm.id = sml.move_id
                left join stock_location as sld on sld.id = sm.location_dest_id
            where sml.state = 'done' and sm.company_id = '%s' and sml.date between '%s' and '%s' 
                and sld.scrap_location = False and sm.picking_id is null
            group by sml.product_id, sml.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from, fields.Datetime.to_datetime(self.ks_date_to)))

        adjusted_date = self.env.cr.fetchall()
        if not adjusted_date:
            return {}
        else:
            ks_dict = dict()
            for ks in adjusted_date: # product_id + location_id + company_id : qty_done(state=done)
                ks_dict[str(ks[0])+str(ks[1])+str(ks[2])] = ks[3]
            adjusted_date = ks_dict
        return adjusted_date


    def ks_data_in_date(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        _select = """ 
            select   product_id, company_id,
                   sum(case when create_date < '{date_from}' then quantity else 0 end) as opening_stock,
                   sum(quantity) as closing_stock,
                    sum(case when create_date >= '{date_from}' then quantity else 0 end) as qty_date,
                   sum(case when create_date >= '{date_from}' and quantity > 0 then quantity else 0 end) as qty_in,
                   sum(case when create_date >= '{date_from}' and quantity < 0 then quantity else 0 end) as qty_out,
            
            
                   sum(case when create_date < '{date_from}' then value else 0 end) as opening_value,
                   sum(value) as closing_value,
                    sum(case when create_date >= '{date_from}' then value else 0 end) as value_date,
                   sum(case when create_date >= '{date_from}' and value > 0 then value else 0 end) as value_in,
                   sum(case when create_date >= '{date_from}' and value < 0 then value else 0 end) as value_out
            from stock_valuation_layer svl
            where create_date <= '{date_to}'
            group by company_id, product_id
            order by product_id
        """
        self.env.cr.execute((_select).format(date_from=ks_date_from, date_to=datetime.strftime(self.ks_date_to, '%Y-%m-%d 23:59:00')))

        dates_in = self.env.cr.fetchall()
        if not dates_in:
            raise ValidationError(_("Opps! There are no data."))
        return dates_in


    def ks_merge_data(self, datas, dates_in, adjusted={}, scrap={}):
        ks_list = []
        kr = self.ks_report
        kid = self.kr_in_dates
        for date in dates_in:
            for data in datas:
                dp_id, dc_id= data[kr['product_id']],  data[kr['company_id']]
                if dp_id == date[kid['product_id']]:
                    ks_cost = self.env['product.product'].browse(date[kid['product_id']]).product_tmpl_id.standard_price
                    ks_adjusted = adjusted.get(str(dp_id)+str(dc_id), 0)
                    ks_scrap = scrap.get(str(dp_id)+str(dc_id), 0)
                    ks_qty_available = date[kid['qty_date']]
                    # if not self.ks_show_exhausted:
                    #     if ks_qty_available < 0: ks_qty_available = 'not_allowed'
                    if ks_qty_available != 'not_allowed':
                        ks_data = 0
                        for rec in ks_list:
                            if rec[0] == data[kr['product_id']]:
                                ks_data = 1
                        if not ks_data:
                            ks_list.append(
                                (data[kr['product_code']], data[kr['product_type']], data[kr['product_categ_id']],
                                 data[kr['product_name']],
                                 date[kid['closing_stock']], ks_cost, data[kr['product_sales_price']],
                                 date[kid['opening_stock']],
                                 date[kid['opening_value']],
                                 date[kid['qty_in']],
                                 date[kid['value_in']],
                                 date[kid['qty_out']],
                                 date[kid['value_out']] ,
                                 date[kid['qty_date']],
                                 date[kid['value_date']] ,
                                 ks_scrap, ks_scrap * ks_cost,
                                 date[kid['closing_stock']],
                                 date[kid['closing_value']], data[kr['product_barcode']]

                                 )
                            )
        if not ks_list:
            raise ValidationError(_("Opps! There are no data."))
        return ks_list


    # def ks_apply_filter(self, data):
    #     ks_data = self.ks_exhausted_filter(data)
    #     if not ks_data:
    #         raise ValidationError(_("Opps! There are no data."))
    #     return ks_data


    # def ks_exhausted_filter(self, data):
    #     ks_data = []
    #     if not self.ks_show_exhausted:
    #         for ks in data:
    #             if ks[6] > 0:  # Ask if product type stockable filter is to be used?
    #                 ks_data.append(ks)
    #     else:
    #         ks_data = data
    #     return ks_data

class KSWarehouseReportValuationOUT(models.AbstractModel):
    _name = "report.ks_warehouse_report.ks_action_warehouse_report_valuation"
    _inherit = 'report.report_xlsx.abstract'

    _description = "Stock Valuation report Out"

    def generate_xlsx_report(self, workbook, data, obj):
        report_name = obj.ks_name
        # One sheet by partner
        bold = workbook.add_format(
            {'font_name': 'Times New Roman', 'bold': True, })
        sheet = workbook.add_worksheet(report_name)
        self.ks_create_workbook_header(report_name, sheet, obj, bold)


        level_1_style = workbook.add_format(
            {'font_name': 'Times New Roman',  'font_size': 10,  'num_format': '#,##0'})
        # get qty available

        self.env.cr.execute("""select      
                                pp.default_code as ks_product_code,
                                pt.type as ks_product_type,
                                pc.id as ks_product_categ_id,
                                pt.name as ks_product_name,
                                rc.id as ks_company_id,
                                pc.name as ks_category,
                                pt.list_price as ks_product_sales_price,
                                sum(svl.quantity) as ks_product_qty_available,
                                pp.id as ks_product_id, 
                                pp.barcode as ks_product_barcode
                                from stock_valuation_layer svl
                                inner join product_product as pp on svl.product_id = pp.id
                                 LEFT JOIN product_template as pt ON pt.id = pp.product_tmpl_id
                                 LEFT JOIN product_category as pc ON pc.id = pt.categ_id
                                 LEFT JOIN res_company as rc ON rc.id = pt.company_id
                                group by pp.id, pt.name, pp.default_code, pp.barcode, pt.type, pc.id, pc.name, pt.list_price, rc.id
                                order by pp.default_code
                         """ )

        datas = self.env.cr.fetchall()
        if not datas:
            raise ValidationError(_("Opps! There are no data."))
        dates_in = obj.ks_data_in_date()

        # ks_adjusted_stock = obj.ks_adjusted_stock()
        #
        # ks_scrap_stock = obj.ks_scrap_stock()

        datas = obj.ks_merge_data(datas, dates_in, )
        if datas:
            i = 1; row = 10; col = 0
            for data in datas:
                # if (not self.location_id or self.location.id == data[4]) or (not self.location_id or self.location.id == data[4])
                sheet.write(row, 0, i)
                sheet.write(row, 1, data[0])
                sheet.write(row, 2, data[19])
                if data[1] == 'product':
                    sheet.write(row, 3, 'Stockable')
                elif data[1]  == 'consu':
                    sheet.write(row, 3, 'Consumable')
                if data[2]:
                    catge_id = self.env['product.category'].browse(int(data[2]))
                sheet.write(row, 4, catge_id.name)
                sheet.write(row, 5, data[3])
                # if data[4]:
                #     location_id = self.env['stock.location'].browse(int(data[4]))
                # sheet.write(row, 6, location_id.display_name)
                # if data[5]:
                #     comp_id = self.env['res.company'].browse(int(data[5]))
                # sheet.write(row, 7, comp_id.name)
                sheet.write(row, 6, data[4])
                sheet.write(row, 7, data[5], level_1_style)
                sheet.write(row, 8, data[6], level_1_style)
                c_1, c_2 = 9, 10
                if obj.ks_show_opening:
                    sheet.write(row, c_1, data[7],level_1_style)
                    sheet.write(row, c_2, data[8], level_1_style)
                    c_1, c_2 = c_1 + 2, c_2 + 2
                    sheet.write(row, c_1, data[9], level_1_style)
                    sheet.write(row, c_2, data[10], level_1_style)
                    c_1, c_2 = c_1 + 2, c_2 + 2
                    sheet.write(row, c_1, data[11], level_1_style)
                    sheet.write(row, c_2, data[12], level_1_style)
                    c_1, c_2 = c_1 + 2, c_2 + 2
                if obj.ks_show_closing:
                    sheet.write(row, c_1, data[13], level_1_style)
                    sheet.write(row, c_2, data[14], level_1_style)
                    c_1, c_2 = c_1 + 2, c_2 + 2
                # if obj.ks_show_adjustment:
                #     sheet.write(row, c_1, data[15], level_1_style)
                #     sheet.write(row, c_2, data[16], level_1_style)
                #     c_1, c_2 = c_1 + 2, c_2 + 2
                # if obj.ks_show_scrap_loss:
                #     sheet.write(row, c_1, data[17], level_1_style)
                #     sheet.write(row, c_2, data[18], level_1_style)
                #     c_1, c_2 = c_1 + 2, c_2 + 2
                if obj.ks_show_current:
                    sheet.write(row, c_1, data[17], level_1_style)
                    sheet.write(row, c_2, data[18], level_1_style)

                row += 1
                i += 1



    # datas = fields.Binary('File', readonly=True)
    # report_name = fields.Char('Report Name', readonly=True)
    def ks_dynamic_sheet(self, sheet, obj, style):
        c_1, c_2 = 9, 10
        if obj.ks_show_opening:
            sheet.write(8, c_1, "Qty.", style)
            sheet.write(8, c_2, "Value", style)
            sheet.write(7, c_1, "Opening Stock", style)
            # self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            # sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2
            sheet.write(8, c_1, "Qty.", style)
            sheet.write(8, c_2, "Value", style)
            sheet.write(7, c_1, "Qty In", style)
            # self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            # sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2
            sheet.write(8, c_1, "Qty.", style)
            sheet.write(8, c_2, "Value", style)
            sheet.write(7, c_1, "Qty Out", style)
            # self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            # sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2

        if obj.ks_show_closing:
            sheet.write(8, c_1, "Qty.", style)
            sheet.write(8, c_2, "Value", style)
            sheet.write(7, c_1, "Closing Stock", style)
            # self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            # sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
            c_1, c_2 = c_1 + 2, c_2 + 2
        # if obj.ks_show_adjustment:
        #     sheet.write(8, c_1, "Qty.", style)
        #     sheet.write(8, c_2, "Value", style)
        #     sheet.write(7, c_1, "Adjustment Stock", style)
        #     # self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
        #     # sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
        #     c_1, c_2 = c_1 + 2, c_2 + 2
        # if obj.ks_show_scrap_loss:
        #     sheet.write(8, c_1, "Qty.", style)
        #     sheet.write(8, c_2, "Value", style)
        #     sheet.write(7, c_1, "Scrap/Loss Stock", style)
        #     # self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
        #     # sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)
        #     c_1, c_2 = c_1 + 2, c_2 + 2
        if obj.ks_show_current:
            sheet.write(8, c_1, "Qty.", style)
            sheet.write(8, c_2, "Value", style)
            sheet.write(7, c_1, "Stock In Hand", style)
            # self.ks_apply_style(sheet.cell(8, c_1), True, True, False, True)
            # sheet.merge_cells(start_row=8, end_row=8, start_column=c_1, end_column=c_2)


    def ks_create_workbook_header(self, report_name, sheet, obj, style):

        sheet.title = str(report_name)

        sheet.merge_range('A1:A2', str(report_name))

        sheet.write('A3',  "COMPANY : " + obj.ks_company_id.name)

        sheet.write('A4', 'FROM : ' + str(obj.ks_date_from) + ' | TO : ' + str(obj.ks_date_to))

        # sheet['A6'] = "REPORT"
        # self.ks_apply_style(sheet['A6'], True, True, 14, True)
        # sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)
        #
        # sheet['A8'] = "S.NO"
        sheet.write('A8',  "S.NO", style)
        sheet.write('B8',  "Reference/Code", style)
        sheet.write('C8',  "Barcode", style)
        sheet.write('D8',  "Type", style)
        sheet.write('E8',  "Category", style)
        sheet.write('F8',  "Product", style)
        sheet.write('G8',  "Available Qty", style)
        sheet.write('H8',  "Cost", style)
        sheet.write('I8',  "Sales Price", style)
        #
        # sheet['B8'] = "Reference/Code"
        # self.ks_apply_style(sheet['B8'], True, True, False, True)
        # sheet.merge_cells(start_row=8, end_row=9, start_column=2, end_column=2)
        #
        # sheet['C8'] = "Barcode"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=3, end_column=3)
        #
        # sheet['D8'] = "Type"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=4, end_column=4)
        #
        # sheet['E8'] = "Category"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=5, end_column=5)
        #
        # sheet['F8'] = "Product"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)
        #
        # sheet['G8'] = "Location"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=7, end_column=7)
        #
        # sheet['H8'] = "Company"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=8, end_column=8)
        #
        # sheet['I8'] = "Available Qty"
        # self.ks_apply_style(sheet['H8'], True, True, False, True)
        # sheet.merge_cells(start_row=8, end_row=9, start_column=9, end_column=9)
        #
        # sheet['J8'] = "Cost"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=10, end_column=10)
        #
        # sheet['K8'] = "Sales Price"
        # sheet.merge_cells(start_row=8, end_row=9, start_column=11, end_column=11)
        #
        self.ks_dynamic_sheet(sheet, obj, style)
        #
        # sheet.freeze_panes = 'C10'
